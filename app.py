import streamlit as st
from streamlit_option_menu import option_menu
import clipboard,os
from openpyxl import workbook,load_workbook

st.set_page_config("Recursion with Wan",layout='centered',initial_sidebar_state="auto",page_icon=":desktop_computer:")

content = {}
allcode = "#include <stdio.h>\n\n"
wb = load_workbook('Book.xlsx')
ws = wb.active

selectedSettings = None
def get_header(get):
    with open("recursivePrograms_byJosuan.h","w") as file:
        file.write(allcode)

for row in ws.iter_rows(min_row=2,values_only=True):
    a_value,b_value,c_value,d_value,e_value = row[0],row[1],row[2],row[3],row[4]
    allcode += f"{b_value}\n\n"
    content[a_value] = [b_value,c_value,d_value,e_value]

    get_header(allcode)

with st.sidebar:
    sidebarOption = ["Download the (.h) file","Copy all the code at ones"]
    selected = st.selectbox("Select",options=sidebarOption)
    if selected == sidebarOption[0]:
        get_header(all)
        with open("recursivePrograms_byJosuan.h", "r") as f:
            filecontents = f.read()
        if st.download_button(
            label="Download",
            data=filecontents,
            file_name="recursivePrograms_byJosuan.h",
            mime="text/plain"
        ):
            st.success("(.h) file successfully downloaded")
    elif selected == sidebarOption[1]:
        if st.button("Copy",key="all_copy"):
            clipboard.copy(allcode)
            st.success("All code copied to clip board")
    st.write("---")

st.header("Welcome to my website where you can find a collection of recursive programs written in C")
st.write("I've got you covered with concise and efficient solutions. Explore the code snippets and enhance your programming skills.")
st.write("---")
for key,value in content.items():
    st.subheader(key.capitalize())
    st.code(value[0],line_numbers=True)
    if st.button("Copy code",key=f"{key}_keys"):
        clipboard.copy(value[0])
        st.success("Code copied to clip board")
    col1,col2 = st.columns(2)
    col1.write(value[1])
    col2.text(value[2])
    st.write(f"Programmed by: {value[3].capitalize()}")
    st.write("---")

for a,b in content.items():
    print(f"{a}\n{b[0]}\n{b[1]}\n{b[2]}\n{b[3]}\n\n");

st.markdown(f'''<p style="text-align:center">© 2023 Josuan. All rights reserved.</p>''',
            unsafe_allow_html=True)

def hide_footer():
    hide_streamlit_style = """
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    </style>
    """
    st.markdown(hide_streamlit_style,unsafe_allow_html=True)
hide_footer()